#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON 데이터를 기본 형태 엑셀로 변환

사용법:
    python basic_excel_generator.py '{"data": [...]}' "output.xlsx"
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 색상 정의 (원본 현황표와 동일)
COLOR_MAP = {
    'YELLOW': 'FFFF99',
    'GREEN': 'C6EFCE',
    'PINK': 'FFCCFF',
    'WHITE': 'FFFFFF'
}


def create_basic_excel(json_data, output_file):
    """
    JSON 데이터를 기본 형태 엑셀로 변환

    Args:
        json_data: {"header": {...}, "data": [...]}
        output_file: 출력 엑셀 파일 경로
    """
    print(f"📂 출력 파일: {output_file}")
    print()

    # JSON 파싱
    if isinstance(json_data, str):
        data = json.loads(json_data)
    else:
        data = json_data

    # 데이터 추출
    if 'data' in data:
        floor_data = data['data']
    else:
        floor_data = data

    if not floor_data:
        print("❌ 데이터가 비어있습니다.")
        return False

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "현황분석"

    # 호수 목록 추출 (첫 번째 층의 units에서)
    first_floor = floor_data[0]
    units = first_floor.get('units', {})
    unit_keys = sorted(units.keys(), key=lambda x: int(''.join(filter(str.isdigit, x)) or '0'))

    print(f"📊 총 {len(unit_keys)}개 호수 감지")
    print(f"📊 총 {len(floor_data)}개 층 감지")
    print()

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더 행 작성
    ws.cell(1, 1, "층")
    for idx, unit_key in enumerate(unit_keys, start=2):
        ws.cell(1, idx, unit_key)

    # 헤더 스타일
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    header_font = Font(bold=True)
    for col_idx in range(1, len(unit_keys) + 2):
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 데이터 행 작성
    print("🔄 데이터 변환 중...")
    for row_idx, floor in enumerate(floor_data, start=2):
        # 층 번호
        floor_text = floor.get('floor', '')
        ws.cell(row_idx, 1, floor_text)
        ws.cell(row_idx, 1).border = thin_border
        ws.cell(row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

        # 각 호수 데이터
        units = floor.get('units', {})
        for col_idx, unit_key in enumerate(unit_keys, start=2):
            unit_data_raw = units.get(unit_key, '')

            # 데이터가 딕셔너리인 경우 처리
            if isinstance(unit_data_raw, dict):
                unit_text = unit_data_raw.get('text', '')
                unit_color = unit_data_raw.get('color', 'WHITE')
            else:
                unit_text = str(unit_data_raw) if unit_data_raw else ''
                unit_color = 'WHITE'

            # 셀에 데이터 쓰기
            cell = ws.cell(row_idx, col_idx, unit_text)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 배경색 적용
            color_key = str(unit_color).upper()
            if color_key in COLOR_MAP:
                cell.fill = PatternFill(start_color=COLOR_MAP[color_key],
                                      end_color=COLOR_MAP[color_key],
                                      fill_type='solid')

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    for col_idx in range(2, len(unit_keys) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12

    # 저장
    print("💾 파일 저장 중...")
    wb.save(output_file)

    print()
    print("=" * 50)
    print("✅ 변환 완료!")
    print(f"📊 총 {len(floor_data)}개 층")
    print(f"📊 총 {len(unit_keys)}개 호수")
    print(f"📂 저장 위치: {output_file}")
    print("=" * 50)

    return True


def main():
    """메인 함수"""
    if len(sys.argv) < 3:
        print("사용법: python basic_excel_generator.py '<JSON 데이터>' <출력파일.xlsx>")
        sys.exit(1)

    json_str = sys.argv[1]
    output_file = sys.argv[2]

    try:
        success = create_basic_excel(json_str, output_file)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
