#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON 데이터를 층호수 형태 엑셀로 변환

사용법:
    python json_to_floor_unit.py '{"header": {...}, "data": [...]}' "output.xlsx"
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 텍스트 색상 정의
TEXT_COLORS = {
    'KT': 'FF0000',  # 빨간색
    'M': '00B050'    # 녹색
}


def extract_floor_number(floor_text):
    """
    "25층" → "25"
    "1층" → "1"
    층 번호 추출
    """
    if not floor_text:
        return ''

    text = str(floor_text).strip()
    # "층" 제거
    if '층' in text:
        text = text.replace('층', '')

    # 숫자만 추출
    return ''.join(filter(str.isdigit, text))


def extract_unit_number(unit_text):
    """
    "1호" → "01"
    "10호" → "10"
    호수 번호 추출 (2자리로 패딩)
    """
    if not unit_text:
        return ''

    text = str(unit_text).strip()
    # "호" 제거
    if '호' in text:
        text = text.replace('호', '')

    # 숫자만 추출
    number = ''.join(filter(str.isdigit, text))

    # 2자리로 패딩
    if number:
        return number.zfill(2)
    return ''


def convert_json_to_floor_unit(json_data, output_file):
    """
    JSON 데이터를 층호수 형태 엑셀로 변환

    Args:
        json_data: {"header": {"building": "...", "name": "..."}, "data": [...]}
        output_file: 출력 엑셀 파일 경로
    """
    print(f"📂 출력 파일: {output_file}")
    print()

    # JSON 파싱
    if isinstance(json_data, str):
        data = json.loads(json_data)
    else:
        data = json_data

    # 데이터 추출
    if 'data' in data:
        # v3 포맷: {header: {...}, data: [...]}
        floor_data = data['data']
    else:
        # 기존 배열 포맷
        floor_data = data

    if not floor_data:
        print("❌ 데이터가 비어있습니다.")
        return False

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "변환결과"

    # 호수 목록 추출 (첫 번째 층의 units에서)
    first_floor = floor_data[0]
    units = first_floor.get('units', {})
    unit_numbers = []

    for unit_key in sorted(units.keys()):
        unit_num = extract_unit_number(unit_key)
        unit_numbers.append((unit_key, unit_num))

    print(f"📊 총 {len(unit_numbers)}개 호수 감지")
    print(f"📊 총 {len(floor_data)}개 층 감지")
    print()

    # 헤더 행 작성 (각 호수마다 2개 열: 층호수 + 데이터)
    ws.cell(1, 1, "층")

    col_idx = 2
    for _, unit_num in unit_numbers:
        if unit_num:
            # 호수 헤더를 2개 열에 걸쳐 병합
            header_text = f"{unit_num}호"
            ws.cell(1, col_idx, header_text)
            # 다음 열은 비워두고 병합
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
            col_idx += 2
        else:
            col_idx += 2

    # 헤더 스타일 적용
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 층 열 + (호수 개수 × 2) 열
    total_cols = 1 + len(unit_numbers) * 2
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 데이터 행 변환
    kt_count = 0
    m_count = 0

    print("🔄 데이터 변환 중...")
    for idx, floor in enumerate(floor_data, start=2):
        # 층 번호 추출
        floor_num = extract_floor_number(floor.get('floor', ''))

        if not floor_num:
            continue

        # 층 번호 작성 (첫 번째 열)
        floor_cell = ws.cell(idx, 1, floor_num)
        floor_cell.border = thin_border
        floor_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 각 호수 처리 (호수마다 2개 열 사용)
        col_idx = 2
        units = floor.get('units', {})

        for unit_key, unit_num in unit_numbers:
            # 해당 호수의 데이터 가져오기
            unit_data_raw = units.get(unit_key, '')

            # 데이터가 딕셔너리인 경우 text 필드 추출
            if isinstance(unit_data_raw, dict):
                unit_data = unit_data_raw.get('text', '')
            else:
                unit_data = str(unit_data_raw) if unit_data_raw else ''

            if unit_num:
                # 층호수 조합
                floor_unit = floor_num + unit_num

                # 첫 번째 열: 층호수 (검정색)
                unit_cell = ws.cell(idx, col_idx, floor_unit)
                unit_cell.font = Font(color='000000')  # 검정색
                unit_cell.border = thin_border
                unit_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 두 번째 열: 데이터 (KT=빨강, M=녹색)
                data_cell = ws.cell(idx, col_idx + 1, unit_data)
                if unit_data in TEXT_COLORS:
                    data_cell.font = Font(color=TEXT_COLORS[unit_data])
                    if unit_data == 'KT':
                        kt_count += 1
                    elif unit_data == 'M':
                        m_count += 1
                data_cell.border = thin_border
                data_cell.alignment = Alignment(horizontal='center', vertical='center')

                col_idx += 2
            else:
                # 호수가 없는 경우 빈 칸 2개
                ws.cell(idx, col_idx).border = thin_border
                ws.cell(idx, col_idx + 1).border = thin_border
                col_idx += 2

    # 열 너비 자동 조정
    # 층 열
    ws.column_dimensions['A'].width = 8

    # 각 호수마다 2개 열 (층호수 열 + 데이터 열)
    for col_idx in range(2, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        if (col_idx - 2) % 2 == 0:
            # 층호수 열
            ws.column_dimensions[col_letter].width = 10
        else:
            # 데이터 열 (KT, M)
            ws.column_dimensions[col_letter].width = 6

    # 저장
    print("💾 파일 저장 중...")
    wb.save(output_file)

    print()
    print("=" * 50)
    print("✅ 변환 완료!")
    print(f"📊 변환된 층: {len(floor_data)}개")
    print(f"🔴 KT: {kt_count}개")
    print(f"🟢 M: {m_count}개")
    print(f"📂 저장 위치: {output_file}")
    print("=" * 50)

    return True


def main():
    """메인 함수"""
    if len(sys.argv) < 3:
        print("사용법: python json_to_floor_unit.py '<JSON 데이터>' <출력파일.xlsx>")
        print()
        print("예시:")
        print('  python json_to_floor_unit.py \'{"data": [...]}\' "output.xlsx"')
        sys.exit(1)

    json_str = sys.argv[1]
    output_file = sys.argv[2]

    try:
        success = convert_json_to_floor_unit(json_str, output_file)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
