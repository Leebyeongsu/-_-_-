import sys
import json
import re
import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# 모든 면 얇은 실선 테두리
THIN_SIDE = Side(style='thin')
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# 텍스트 변환 규칙: 공백 제거 후 키 매칭 → (변환텍스트, 글자색hex)
# 연한 녹색: 70AD47 / 진한 녹색: 375623 / 진한 빨간색: C00000
UNIT_TEXT_TRANSFORM = {
    'M(미동의)':   ('KT', 'C00000'),  # 진한 빨간색 글자
    'M':           ('M',  '375623'),  # 녹색 글자
    'IM(미동의)':  ('KT', 'C00000'),  # 진한 빨간색 글자
    'ITM(미동의)': ('KT', 'C00000'),  # 진한 빨간색 글자
    'IT(미동의)':  ('KT', 'C00000'),
    'IM':          ('KT', 'C00000'),
    'ITM':         ('KT', 'C00000'),
    'TM':          ('KT', 'C00000'),
    'IT':          ('KT', 'C00000'),
    'I':           ('KT', 'C00000'),
    'TM(미동의)':  ('KT', 'C00000'),  # 진한 빨간색 글자
    'I(미동의)':   ('KT', 'C00000'),  # 진한 빨간색 글자
    'T(미동의)':   ('',   None),       # 빈 셀
}
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def find_buildings(ws):
    """행 1~3에서 동 이름과 컬럼 범위 추출 (병합 셀 기준)"""
    buildings = []
    found_names = set()

    # 1. 병합된 셀 탐색 (행 1~3 범위)
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row <= 3:
            cell_val = ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
            if cell_val:
                val_str = str(cell_val).strip()
                if re.search(r'\d+동', val_str) and val_str not in found_names:
                    found_names.add(val_str)
                    buildings.append({
                        'name': val_str,
                        'col_start': merge_range.min_col,
                        'col_end': merge_range.max_col,
                    })

    # 2. 병합 없는 일반 셀 탐색 (fallback)
    merged_coords = set()
    for merge_range in ws.merged_cells.ranges:
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merged_coords.add((r, c))

    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            if (row, col) in merged_coords:
                continue
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                val_str = str(cell_val).strip()
                if re.search(r'\d+동', val_str) and val_str not in found_names:
                    found_names.add(val_str)
                    buildings.append({
                        'name': val_str,
                        'col_start': col,
                        'col_end': col,
                    })

    buildings.sort(key=lambda b: b['col_start'])
    return buildings


def copy_cell(src, dst):
    """셀 값 + 서식(색상, 폰트, 테두리, 정렬) 복사"""
    dst.value = src.value
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def move_unit_subdata(new_ws, num_src_cols):
    """전체 시트에서 호수 행을 탐색:
    - 바로 아래 데이터를 호수 옆 빈 열로 이동 (텍스트만, 배경색 제외)
    - 이동 후 데이터 행 삭제
    """
    unit_pattern = re.compile(r'^\d+호?$')
    rows_to_delete = []

    for row in range(1, new_ws.max_row + 1):
        is_unit_row = any(
            new_ws.cell(row=row, column=i * 2 + 1).value and
            unit_pattern.match(str(new_ws.cell(row=row, column=i * 2 + 1).value).strip())
            for i in range(num_src_cols)
        )
        if not is_unit_row:
            continue

        data_row = row + 1
        if data_row > new_ws.max_row:
            continue

        rows_to_delete.append(data_row)

        for i in range(num_src_cols):
            odd_col = i * 2 + 1   # 호수 열
            even_col = i * 2 + 2  # 호수 옆 빈 열

            src_cell = new_ws.cell(row=data_row, column=odd_col)
            if src_cell.value is not None and str(src_cell.value).strip() != '':
                dst_cell = new_ws.cell(row=row, column=even_col)
                # 공백 모두 제거 후 변환 규칙 적용
                key = re.sub(r'\s+', '', str(src_cell.value))
                if key in UNIT_TEXT_TRANSFORM:
                    new_text, fg_hex = UNIT_TEXT_TRANSFORM[key]
                    dst_cell.value = new_text if new_text != '' else None
                    if fg_hex:
                        dst_cell.font = Font(color=fg_hex, bold=True)
                else:
                    dst_cell.value = src_cell.value
                dst_cell.alignment = LEFT_ALIGN
                dst_cell.border = ALL_BORDER

    # 데이터 행 역순 삭제 (행 번호 밀림 방지)
    for del_row in sorted(set(rows_to_delete), reverse=True):
        new_ws.delete_rows(del_row)


def create_building_sheets(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active

    buildings = find_buildings(ws)

    if not buildings:
        return {
            'success': False,
            'error': '1~3행에서 동 이름(예: 101동)을 찾을 수 없습니다.'
        }

    ws.title = '원본'

    created_sheets = []

    for building in buildings:
        sheet_name = re.sub(r'[\\/*?:\[\]]', '_', building['name'])[:31]
        if sheet_name in wb.sheetnames:
            continue

        col_start = building['col_start']
        col_end = building['col_end']

        new_ws = wb.create_sheet(title=sheet_name)

        # ── 열 너비 설정 ──
        # 홀수 열(호수 열): 5.00 고정 / 짝수 열(빈 열): 원본과 동일
        for i, src_col_idx in enumerate(range(col_start, col_end + 1)):
            src_letter = get_column_letter(src_col_idx)
            dst_data_letter = get_column_letter(i * 2 + 1)
            dst_blank_letter = get_column_letter(i * 2 + 2)

            src_width = ws.column_dimensions[src_letter].width \
                if src_letter in ws.column_dimensions else 8.43
            new_ws.column_dimensions[dst_data_letter].width = 5.00   # 호수 열 고정
            new_ws.column_dimensions[dst_blank_letter].width = src_width

        # ── 행 높이: 전체 25 고정 ──
        for row_idx in range(1, ws.max_row + 1):
            new_ws.row_dimensions[row_idx].height = 25

        # ── 셀 데이터 + 서식 복사 (데이터 열: 홀수 위치) ──
        # ── 빈 열(짝수 위치)에 모든 테두리 적용 ──
        num_src_cols = col_end - col_start + 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_start, max_col=col_end):
            for src_cell in row:
                offset = src_cell.column - col_start      # 0-based
                dst_col = offset * 2 + 1                  # 홀수 위치 (데이터)
                dst_cell = new_ws.cell(row=src_cell.row, column=dst_col)
                copy_cell(src_cell, dst_cell)
                # 데이터 셀에도 테두리 강제 적용
                dst_cell.border = ALL_BORDER

                # 오른쪽 빈 열에 테두리 적용
                blank_cell = new_ws.cell(row=src_cell.row, column=dst_col + 1)
                blank_cell.border = ALL_BORDER

        # ── 병합 셀 복사 (빈 열 포함하여 확장) ──
        for merge_range in list(ws.merged_cells.ranges):
            if (merge_range.min_col >= col_start and
                    merge_range.max_col <= col_end):
                min_offset = merge_range.min_col - col_start   # 0-based
                max_offset = merge_range.max_col - col_start   # 0-based

                new_min_col = min_offset * 2 + 1

                if merge_range.min_col < merge_range.max_col:
                    # 가로로 여러 컬럼에 걸친 병합 → 빈 열 포함하여 확장
                    new_max_col = max_offset * 2 + 2
                else:
                    # 세로 병합(단일 열) → 확장 없음
                    new_max_col = max_offset * 2 + 1

                new_ws.merge_cells(
                    start_row=merge_range.min_row,
                    start_column=new_min_col,
                    end_row=merge_range.max_row,
                    end_column=new_max_col
                )

        # ── 호수 아래 데이터 → 호수 옆 빈 열로 이동 (전체 시트) ──
        move_unit_subdata(new_ws, num_src_cols)

        created_sheets.append(sheet_name)

    wb.save(output_path)

    return {
        'success': True,
        'buildings': [b['name'] for b in buildings],
        'created_sheets': created_sheets,
        'total': len(created_sheets)
    }


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({
            'success': False,
            'error': '사용법: python map_converter.py <input.xlsx> <output.xlsx>'
        }, ensure_ascii=False))
        sys.exit(1)

    result = create_building_sheets(sys.argv[1], sys.argv[2])
    print(json.dumps(result, ensure_ascii=False))
