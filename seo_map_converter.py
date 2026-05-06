import sys
import json
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

THIN_SIDE = Side(style='thin')
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

FONT_KT    = Font(color="C00000", bold=True)   # 빨간색
FONT_M     = Font(color="375623", bold=True)   # 녹색
FONT_TITLE = Font(bold=True, size=12)

ALIGN_CENTER      = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT_INDENT = Alignment(horizontal='left',   vertical='center')

GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 변경 2


def to_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def get_status(f_num, e_num):
    if f_num >= 1:
        return "KT", FONT_KT
    if f_num == 0 and e_num >= 1:
        return "M", FONT_M
    return None, None


def get_apt_name(ws):
    """요약 시트 A열 첫 번째 비어있지 않은 문자열 = 아파트 이름"""
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str) and val.strip():
            return val.strip()
    return ""


def collect_data(ws):
    """요약 시트 B=동명, C=호수, E=E값, F=F값 수집"""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        b_val = row[1] if len(row) > 1 else None
        c_val = row[2] if len(row) > 2 else None
        e_val = row[4] if len(row) > 4 else None
        f_val = row[5] if len(row) > 5 else None

        if not b_val or not c_val:
            continue

        b_str = str(b_val).strip()
        c_str = str(c_val).strip()

        if not re.search(r'\d+동', b_str):
            continue

        unit_str = c_str.replace("호", "").strip()
        try:
            unit_num = int(unit_str)
        except ValueError:
            continue

        rows.append({
            "building": b_str,
            "unit_raw": c_str,
            "unit_num": unit_num,
            "floor": unit_num // 100,
            "line":  unit_num % 100,
            "e_num": to_float(e_val),
            "f_num": to_float(f_val),
        })

    return rows


def set_border_range(ws, min_row, max_row, min_col, max_col):
    """지정 범위 모든 셀에 테두리 적용"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = ALL_BORDER


def build_sheet(wb, building, bldg_data, apt_name):
    lines  = sorted(set(r["line"]  for r in bldg_data))
    floors = sorted(set(r["floor"] for r in bldg_data), reverse=True)
    data_map = {(r["floor"], r["line"]): r for r in bldg_data}

    sheet_name = re.sub(r'[\\/*?:\[\]]', '_', building)[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # ── 행 1: 타이틀 (아파트명 + 동명), 4셀 병합 ──────────────── 변경 1
    title_text = f"{apt_name} {building}".strip()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font      = FONT_TITLE
    title_cell.alignment = ALIGN_CENTER
    set_border_range(ws, 1, 1, 1, 4)

    # ── 행 2: 라인 헤더, 두 셀 병합 + 연한 회색 ──────────────── 변경 2
    for idx, line in enumerate(lines):
        col_unit   = idx * 2 + 1
        col_status = idx * 2 + 2

        ws.merge_cells(start_row=2, start_column=col_unit, end_row=2, end_column=col_status)
        hdr = ws.cell(row=2, column=col_unit, value=f"{line}라인")  # 앞 0 제거
        hdr.font      = Font(bold=True)
        hdr.alignment = ALIGN_CENTER
        hdr.fill      = GRAY_FILL
        set_border_range(ws, 2, 2, col_unit, col_status)

    # ── 행 3~: 데이터 (층 내림차순) ──
    for row_idx, floor in enumerate(floors, start=3):
        for idx, line in enumerate(lines):
            col_unit   = idx * 2 + 1
            col_status = idx * 2 + 2

            unit_cell   = ws.cell(row=row_idx, column=col_unit)
            status_cell = ws.cell(row=row_idx, column=col_status)

            unit_cell.border   = ALL_BORDER
            status_cell.border = ALL_BORDER

            entry = data_map.get((floor, line))
            if not entry:
                continue

            unit_cell.value     = entry["unit_raw"]
            unit_cell.alignment = ALIGN_CENTER

            text, font = get_status(entry["f_num"], entry["e_num"])
            if text:
                status_cell.value     = f"  {text}"   # 공백 2칸으로 들여쓰기
                status_cell.font      = font
                status_cell.alignment = ALIGN_LEFT_INDENT

    return sheet_name


def seo_convert(input_path, output_path):
    wb = load_workbook(input_path)

    if "요약" not in wb.sheetnames:
        return {"success": False, "error": '"요약" 시트를 찾을 수 없습니다.'}

    ws_summary = wb["요약"]
    apt_name   = get_apt_name(ws_summary)      # 변경 1: 아파트명 추출

    all_data = collect_data(ws_summary)
    if not all_data:
        return {"success": False, "error": "요약 시트에서 유효한 데이터를 찾을 수 없습니다."}

    buildings = sorted(
        set(r["building"] for r in all_data),
        key=lambda x: int(re.search(r'\d+', x).group())
    )

    created_sheets = []
    for building in buildings:
        bldg_data = [r for r in all_data if r["building"] == building]
        if not bldg_data:
            continue
        sheet_name = build_sheet(wb, building, bldg_data, apt_name)
        created_sheets.append(sheet_name)

    wb.save(output_path)

    return {
        "success": True,
        "buildings": buildings,
        "created_sheets": created_sheets,
        "total": len(created_sheets)
    }


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({
            "success": False,
            "error": "사용법: python seo_map_converter.py <input.xlsx> <output.xlsx>"
        }, ensure_ascii=False))
        sys.exit(1)

    result = seo_convert(sys.argv[1], sys.argv[2])
    print(json.dumps(result, ensure_ascii=False))
