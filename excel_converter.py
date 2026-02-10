


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 변환 스크립트
다운로드한 현황표 엑셀 파일을 "층호수 + 데이터" 형태로 변환

사용법:
    python excel_converter.py "입력파일.xlsx"
    python excel_converter.py "입력파일.xlsx" "출력파일.xlsx"
"""

import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 색상 정의 (원본 엑셀의 배경색 RGB 값)
COLOR_THRESHOLDS = {
    'GREEN': (0xC6, 0xEF, 0xCE),   # C6EFCE - 녹색
    'PINK': (0xFF, 0xCC, 0xFF),     # FFCCFF - 핑크
    'YELLOW': (0xFF, 0xFF, 0x99),   # FFFF99 - 노란색
    'WHITE': (0xFF, 0xFF, 0xFF)     # FFFFFF - 흰색
}

# 텍스트 색상 정의
TEXT_COLORS = {
    'KT': 'FF0000',  # 빨간색
    'M': '00B050'    # 녹색
}


def rgb_to_hex(rgb):
    """RGB 튜플을 HEX 문자열로 변환"""
    if rgb is None:
        return 'FFFFFF'
    return '%02X%02X%02X' % rgb


def hex_to_rgb(hex_str):
    """HEX 문자열을 RGB 튜플로 변환"""
    if hex_str is None or len(hex_str) < 6:
        return (255, 255, 255)

    # ARGB 형식인 경우 (8자리)
    if len(hex_str) == 8:
        hex_str = hex_str[2:]  # 알파 채널 제거

    try:
        return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))
    except:
        return (255, 255, 255)


def color_distance(rgb1, rgb2):
    """두 RGB 색상 간의 유클리드 거리 계산"""
    return sum((a - b) ** 2 for a, b in zip(rgb1, rgb2)) ** 0.5


def identify_color(rgb):
    """
    RGB 값으로부터 색상 유형 판별
    가장 가까운 색상을 반환
    """
    if rgb is None:
        return 'WHITE'

    min_distance = float('inf')
    identified = 'WHITE'

    for color_name, color_rgb in COLOR_THRESHOLDS.items():
        distance = color_distance(rgb, color_rgb)
        if distance < min_distance:
            min_distance = distance
            identified = color_name

    return identified


def get_cell_bg_color(cell):
    """셀의 배경색 RGB 추출"""
    if cell.fill and cell.fill.fgColor:
        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
            hex_color = cell.fill.fgColor.rgb
            # ARGB 형식인 경우 (00RRGGBB)
            if isinstance(hex_color, str):
                return hex_to_rgb(hex_color)
            return (255, 255, 255)
    return (255, 255, 255)


def color_to_data(color_type):
    """
    색상 유형에 따라 데이터 텍스트 반환
    GREEN, PINK → "KT"
    YELLOW → "M"
    WHITE → ""
    """
    if color_type in ['GREEN', 'PINK']:
        return 'KT'
    elif color_type == 'YELLOW':
        return 'M'
    else:
        return ''


def extract_floor_number(floor_text):
    """
    "25층" → "25"
    "1층" → "1"
    층 번호 추출
    """
    if not floor_text:
        return ''

    text = str(floor_text).strip()
    # "층" 제거
    if '층' in text:
        text = text.replace('층', '')

    # 숫자만 추출
    return ''.join(filter(str.isdigit, text))


def extract_unit_number(unit_text):
    """
    "1호" → "01"
    "10호" → "10"
    호수 번호 추출 (2자리로 패딩)
    """
    if not unit_text:
        return ''

    text = str(unit_text).strip()
    # "호" 제거
    if '호' in text:
        text = text.replace('호', '')

    # 숫자만 추출
    number = ''.join(filter(str.isdigit, text))

    # 2자리로 패딩
    if number:
        return number.zfill(2)
    return ''


def convert_excel(input_file, output_file=None):
    """
    엑셀 파일 변환

    Args:
        input_file: 입력 엑셀 파일 경로
        output_file: 출력 엑셀 파일 경로 (None이면 자동 생성)
    """
    if not os.path.exists(input_file):
        print(f"❌ 파일을 찾을 수 없습니다: {input_file}")
        return False

    # 출력 파일명 생성
    if output_file is None:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_변환{ext}"

    print(f"📂 입력 파일: {input_file}")
    print(f"📂 출력 파일: {output_file}")
    print()

    # 엑셀 파일 로드
    print("🔄 엑셀 파일 로딩 중...")
    wb = load_workbook(input_file)
    ws = wb.active

    # 새 워크북 생성
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "변환결과"

    # 헤더 행 읽기 (첫 번째 행)
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(1, col_idx)
        headers.append(cell.value)

    print(f"📊 총 {len(headers)}개 열 감지")
    print(f"📊 총 {ws.max_row - 1}개 층 감지")
    print()

    # 호수 목록 추출 (첫 번째 열 제외)
    unit_numbers = []
    for i, header in enumerate(headers[1:], start=1):
        unit_num = extract_unit_number(header)
        unit_numbers.append((i + 1, unit_num))  # (원본 열 인덱스, 호수)

    # 헤더 행 작성 (각 호수마다 2개 열: 층호수 + 데이터)
    new_ws.cell(1, 1, "층")

    col_idx = 2
    for _, unit_num in unit_numbers:
        if unit_num:
            # 호수 헤더를 2개 열에 걸쳐 병합
            header_text = f"{unit_num}호"
            new_ws.cell(1, col_idx, header_text)
            # 다음 열은 비워두고 병합
            new_ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
            col_idx += 2
        else:
            col_idx += 2

    # 헤더 스타일 적용
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 층 열 + (호수 개수 × 2) 열
    total_cols = 1 + len(unit_numbers) * 2
    for col_idx in range(1, total_cols + 1):
        cell = new_ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 데이터 행 변환
    converted_count = 0
    kt_count = 0
    m_count = 0

    print("🔄 데이터 변환 중...")
    for row_idx in range(2, ws.max_row + 1):
        # 층 번호 추출
        floor_cell = ws.cell(row_idx, 1)
        floor_num = extract_floor_number(floor_cell.value)

        if not floor_num:
            continue

        new_row_idx = row_idx

        # 층 번호 작성 (첫 번째 열)
        floor_cell = new_ws.cell(new_row_idx, 1, floor_num)
        floor_cell.border = thin_border
        floor_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 각 호수 처리 (호수마다 2개 열 사용)
        new_col_idx = 2
        for orig_col_idx, unit_num in unit_numbers:
            # 원본 셀 읽기
            orig_cell = ws.cell(row_idx, orig_col_idx)
            bg_color_rgb = get_cell_bg_color(orig_cell)
            color_type = identify_color(bg_color_rgb)

            # 데이터 생성
            data_text = color_to_data(color_type)

            # 층호수 조합
            if unit_num:
                floor_unit = floor_num + unit_num

                # 첫 번째 열: 층호수 (검정색)
                unit_cell = new_ws.cell(new_row_idx, new_col_idx, floor_unit)
                unit_cell.font = Font(color='000000')  # 검정색
                unit_cell.border = thin_border
                unit_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 두 번째 열: 데이터 (KT=빨강, M=녹색)
                data_cell = new_ws.cell(new_row_idx, new_col_idx + 1, data_text)
                if data_text in TEXT_COLORS:
                    data_cell.font = Font(color=TEXT_COLORS[data_text])
                    if data_text == 'KT':
                        kt_count += 1
                    elif data_text == 'M':
                        m_count += 1
                data_cell.border = thin_border
                data_cell.alignment = Alignment(horizontal='center', vertical='center')

                new_col_idx += 2
            else:
                # 호수가 없는 경우 빈 칸 2개
                new_ws.cell(new_row_idx, new_col_idx).border = thin_border
                new_ws.cell(new_row_idx, new_col_idx + 1).border = thin_border
                new_col_idx += 2

        converted_count += 1

    # 열 너비 자동 조정
    # 층 열
    new_ws.column_dimensions['A'].width = 8

    # 각 호수마다 2개 열 (층호수 열 + 데이터 열)
    total_cols = 1 + len(unit_numbers) * 2
    for col_idx in range(2, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        if (col_idx - 2) % 2 == 0:
            # 층호수 열
            new_ws.column_dimensions[col_letter].width = 10
        else:
            # 데이터 열 (KT, M)
            new_ws.column_dimensions[col_letter].width = 6

    # 저장
    print("💾 파일 저장 중...")
    new_wb.save(output_file)

    print()
    print("=" * 50)
    print("✅ 변환 완료!")
    print(f"📊 변환된 층: {converted_count}개")
    print(f"🔴 KT: {kt_count}개")
    print(f"🟢 M: {m_count}개")
    print(f"📂 저장 위치: {output_file}")
    print("=" * 50)

    return True


def main():
    """메인 함수"""
    if len(sys.argv) < 2:
        print("사용법: python excel_converter.py <입력파일.xlsx> [출력파일.xlsx]")
        print()
        print("예시:")
        print('  python excel_converter.py "현황표_1234567890.xlsx"')
        print('  python excel_converter.py "현황표.xlsx" "변환결과.xlsx"')
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        success = convert_excel(input_file, output_file)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
